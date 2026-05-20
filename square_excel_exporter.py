"""
square_excel_exporter.py
========================
Excel export for Square POS reports.
One sheet per outlet (plus a Combined sheet).
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

_CURRENCY_COLS = {"Net Sales £", "Gross Sales £"}
_RIGHT_COLS    = {"#SL", "Qty", "Net Sales £", "Gross Sales £", "Total"}
_CENTER_COLS   = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}

def _get_palette(title: str) -> dict:
    t = title.lower()
    if "weekly" in t:
        return {
            "title_bg": "FFFFFF", "title_fg": "000000",
            "header_bg": "FFFFFF", "header_fg": "000000",
            "total_bg": "FFFFFF", "total_fg": "000000",
            "alt_bg": "FFFFFF", "border": "000000"
        }
    elif "best setup" in t or "best" in t or "top 10" in t:
        return {
            "title_bg": "A7F3D0", "title_fg": "064E3B",   # Light green
            "header_bg": "D1FAE5", "header_fg": "065F46",
            "total_bg": "6EE7B7", "total_fg": "064E3B",
            "alt_bg": "F0FDF4", "border": "A7F3D0"
        }
    elif "worst" in t:
        return {
            "title_bg": "FECACA", "title_fg": "7F1D1D",   # Light red
            "header_bg": "FEE2E2", "header_fg": "991B1B",
            "total_bg": "FCA5A5", "total_fg": "7F1D1D",
            "alt_bg": "FEF2F2", "border": "FECACA"
        }
    else:
        return {
            "title_bg": "E9D5FF", "title_fg": "4C1D95",   # Light purple
            "header_bg": "F3E8FF", "header_fg": "581C87",
            "total_bg": "D8B4FE", "total_fg": "3B0764",
            "alt_bg": "FAF5FF", "border": "E9D5FF"
        }

def _side(color):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_block(ws, df: pd.DataFrame, title: str, start_row: int, start_col: int) -> int:
    """Write one report table to worksheet. Returns next free column."""
    n_cols = len(df.columns)

    pal = _get_palette(title)
    is_weekly = "weekly" in title.lower()

    # ── Title bar ────────────────────────────────────────────────────────────
    tc = ws.cell(row=start_row, column=start_col, value=title)
    tc.font  = Font(bold=True, color=pal["title_fg"], size=10)
    tc.fill  = PatternFill("solid", fgColor=pal["title_bg"])
    
    if is_weekly:
        tc.alignment = Alignment(horizontal="left", vertical="center")
    else:
        tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row,   end_column=start_col + n_cols - 1,
    )
    
    if is_weekly:
        for c in range(start_col, start_col + n_cols):
            b_l = Side(style="medium", color="000000") if c == start_col else None
            b_r = Side(style="medium", color="000000") if c == start_col + n_cols - 1 else None
            ws.cell(row=start_row, column=c).border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="thin", color="000000"), left=b_l, right=b_r)
            
    ws.row_dimensions[start_row].height = 22
    cur_row = start_row + 1

    # ── Header row ───────────────────────────────────────────────────────────
    for ci, col in enumerate(df.columns, start=start_col):
        disp_col = " " if is_weekly and str(col) == "Item Name" else str(col)
        cell = ws.cell(row=cur_row, column=ci, value=disp_col)
        cell.font      = Font(bold=True, color=pal["header_fg"], size=10)
        cell.fill      = PatternFill("solid", fgColor=pal["header_bg"])
        
        if is_weekly:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            b_l = Side(style="medium", color="000000") if ci == start_col else Side(style="thin", color="000000")
            b_r = Side(style="medium", color="000000") if ci == start_col + n_cols - 1 else Side(style="thin", color="000000")
            cell.border = Border(top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"), left=b_l, right=b_r)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = _side(pal["border"])
            
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    # ── Data rows ────────────────────────────────────────────────────────────
    data_rows  = df.iloc[:-1]   # everything except last (total)
    total_row  = df.iloc[-1]

    last_ri = len(data_rows) - 1
    for ri, (_, row) in enumerate(data_rows.iterrows()):
        is_subtotal = is_weekly and str(row.iloc[0]).endswith("Total") and str(row.iloc[0]) != "Total"
        is_grand_total = is_weekly and str(row.iloc[0]) == "Total"
        is_blank = is_weekly and str(row.iloc[0]).strip() == ""
        is_last_row = (ri == last_ri)
        
        bg = pal["alt_bg"] if (ri % 2 == 1 and not is_subtotal and not is_grand_total and not is_blank) else "FFFFFF"
        
        for ci, (col, val) in enumerate(row.items(), start=start_col):
            disp_val = "" if is_blank else val
            cell = ws.cell(row=cur_row, column=ci, value=disp_val)
            
            if is_weekly:
                is_label_col = (ci == start_col)
                is_last_col = (ci == start_col + n_cols - 1)
                is_penultimate_col = (ci == start_col + n_cols - 2)
                is_col_total = (col == "Total")
                
                is_lower_block = (is_blank or is_subtotal or is_grand_total)
                
                if is_lower_block:
                    b_l, b_r, b_t, b_b = None, None, None, None
                    
                    if is_blank:
                        # Gap Row: only extreme outer sides. No inner column dropping.
                        if is_label_col: b_l = Side(style="medium", color="000000")
                        if is_last_col:  b_r = Side(style="medium", color="000000")
                    else:
                        # Subtotal / Grand Total blocks
                        if is_label_col: b_l = Side(style="medium", color="000000")
                        if is_last_col:  b_r = Side(style="medium", color="000000")
                        
                        # Total column separator
                        if is_col_total: b_l = Side(style="medium", color="000000")
                        if is_penultimate_col: b_r = Side(style="medium", color="000000")
                        
                        if is_subtotal:
                            is_last_subtotal = (ri == len(data_rows) - 2)
                            is_first_subtotal = str(data_rows.iloc[max(0, ri-1)].iloc[0]).strip() == ""
                            
                            # Complete top horizon over data
                            if is_first_subtotal and not is_label_col:
                                b_t = Side(style="medium", color="000000")
                            
                            # Complete bottom horizon over data
                            if is_last_subtotal and not is_label_col:
                                b_b = Side(style="medium", color="000000")
                                
                    cell.border = Border(left=b_l, right=b_r, top=b_t, bottom=b_b)
                
                else: # Regular item row
                    b_l = Side(style="medium", color="000000") if is_label_col else Side(style="thin", color="000000")
                    b_r = Side(style="medium", color="000000") if (is_last_col or is_penultimate_col) else Side(style="thin", color="000000")
                    b_t = Side(style="thin", color="000000")
                    
                    # Bottom border is medium if the next row is the gap row!
                    is_next_gap = (ri + 1 < len(data_rows)) and str(data_rows.iloc[ri+1].iloc[0]).strip() == ""
                    b_b = Side(style="medium", color="000000") if is_next_gap else Side(style="thin", color="000000")
                    
                    cell.border = Border(left=b_l, right=b_r, top=b_t, bottom=b_b)
                    
            else:
                if is_blank:
                    continue
                cell.border = _side(pal["border"])
                
            cell.fill = PatternFill("solid", fgColor=bg)
            
            is_bold = is_subtotal or is_grand_total or (is_weekly and col == "Total")
            if is_bold:
                cell.font = Font(bold=True, color="000000")
            
            if is_weekly and col in _CENTER_COLS:
                halign = "center"
            elif col in _RIGHT_COLS:
                halign = "right"
            else:
                halign = "right" if (is_subtotal or is_grand_total) and ci == start_col else "left"
                
            cell.alignment = Alignment(horizontal=halign)
            
            if col in _CURRENCY_COLS and isinstance(val, (int, float)):
                cell.number_format = '"£"#,##0.00'
        ws.row_dimensions[cur_row].height = 15
        cur_row += 1

    # ── Total row ────────────────────────────────────────────────────────────
    is_grand_total = is_weekly  # In weekly report, total_row is the Grand Total
    for ci, (col, val) in enumerate(total_row.items(), start=start_col):
        cell = ws.cell(row=cur_row, column=ci, value=val)
        cell.font      = Font(bold=True, color=pal["total_fg"])
        cell.fill      = PatternFill("solid", fgColor=pal["total_bg"])
        
        if is_weekly:
            is_col_total = (col == "Total")
            is_label_col = (ci == start_col)
            is_last_col  = (ci == start_col + n_cols - 1)
            is_penultimate_col = (ci == start_col + n_cols - 2)
            
            b_l = None; b_r = None; b_t = None
            if is_label_col: b_l = Side(style="medium", color="000000")
            if is_last_col:  b_r = Side(style="medium", color="000000")
            
            if is_col_total: b_l = Side(style="medium", color="000000")
            if is_penultimate_col: b_r = Side(style="medium", color="000000")
            
            b_b = Side(style="medium", color="000000")
            
            cell.border = Border(top=b_t, bottom=b_b, left=b_l, right=b_r)
            halign = "right" if is_label_col else ("center" if col in _CENTER_COLS else "right")
            cell.alignment = Alignment(horizontal=halign)
            
        else:
            cell.border    = _side(pal["header_bg"])
            cell.alignment = Alignment(
                horizontal="right" if col in _RIGHT_COLS else "left"
            )
            
        if col in _CURRENCY_COLS and isinstance(val, (int, float)):
            cell.number_format = '"£"#,##0.00'
    ws.row_dimensions[cur_row].height = 18

    # ── Column widths ────────────────────────────────────────────────────────
    for offset, col in enumerate(df.columns):
        ci = start_col + offset
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        col_letter = get_column_letter(ci)
        curr_width = ws.column_dimensions[col_letter].width
        if curr_width is None:
            curr_width = 8.0
        new_width = min(max_len + 4, 45)
        ws.column_dimensions[col_letter].width = max(curr_width, new_width)

    return start_col + n_cols + 1   # gap column


def build_square_excel(reports: dict[str, tuple[pd.DataFrame, str]]) -> bytes:
    """
    Build an Excel workbook placing all reports vertically on a single sheet.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Square Report"
    ws.sheet_view.showGridLines = False

    ws.freeze_panes = "A3"

    ordered = sorted(
        reports.items(),
        key=lambda kv: (0 if "Combined" in kv[0] else 1, kv[0]),
    )

    start_col = 1
    for label, (df, title) in ordered:
        start_col = _write_block(ws, df, title, start_row=1, start_col=start_col)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

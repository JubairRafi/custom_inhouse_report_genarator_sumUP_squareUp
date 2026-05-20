import openpyxl, json, os, datetime

BASE = r"c:\St george\hand made Tools\weekly sales reports genarator manager\report"

def load_weekly(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    def parse_side(col_off, data_rows):
        title = rows[0][col_off]
        items, total_row = [], None
        for r in data_rows:
            name = r[col_off]
            if name is None: continue
            if str(name).strip() in ("Food - Wine Bar Total", "Total"):
                total_row = r; continue
            vals = [r[col_off+1+i] or 0 for i in range(7)]
            tot  = r[col_off+8] or 0
            items.append({"name": str(name), "days": vals, "total": tot})
        grand       = [total_row[col_off+1+i] or 0 for i in range(7)] if total_row else [0]*7
        grand_total = (total_row[col_off+8] or 0) if total_row else 0
        return {"title": str(title), "items": items, "grand": grand, "grand_total": grand_total}
    return parse_side(0, rows[2:]), parse_side(10, rows[2:])

def load_combined(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    combined = []
    for r in rows[2:]:
        if r[1] and str(r[1]).strip() != "TOTAL":
            combined.append({"name": str(r[1]), "kr": r[2] or 0, "sm": r[3] or 0, "total": r[4] or 0})
    total_kr  = next((r[2] for r in rows if r[1] and str(r[1])=="TOTAL"), 0) or 0
    total_sm  = next((r[3] for r in rows if r[1] and str(r[1])=="TOTAL"), 0) or 0
    total_all = next((r[4] for r in rows if r[1] and str(r[1])=="TOTAL"), 0) or 0
    return combined, int(total_kr), int(total_sm), int(total_all)

wk1_sm, wk1_kr = load_weekly(os.path.join(BASE, "Square_Weekly_Kings_Road_(04-05-26-10-05-26).xlsx"))
wk2_sm, wk2_kr = load_weekly(os.path.join(BASE, "Square_Weekly_Kings_Road_(11-05-26-17-05-26).xlsx"))
combined, total_kr, total_sm, total_all = load_combined(
    os.path.join(BASE, "Square_All_Items_Comb._(04-05-26-17-05-26) (1).xlsx"))

days  = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
top12 = [x for x in combined if "Annul" not in x["name"]][:12]
wk1_daily = [wk1_sm["grand"][i]+wk1_kr["grand"][i] for i in range(7)]
wk2_daily = [wk2_sm["grand"][i]+wk2_kr["grand"][i] for i in range(7)]

def rows_html(items):
    out = []
    for x in items:
        cls = "annule" if "Annul" in x["name"] else ""
        cells = "".join(f'<td class="num">{v if v else ""}</td>' for v in x["days"])
        out.append(f'<tr class="{cls}"><td>{x["name"]}</td>{cells}<td class="num"><strong>{x["total"]}</strong></td></tr>')
    return "\n".join(out)

def total_row_html(g, gt):
    cells = "".join(f'<td class="num"><strong>{v}</strong></td>' for v in g)
    return f'<tr class="total-row"><td><strong>TOTAL</strong></td>{cells}<td class="num"><strong>{gt}</strong></td></tr>'

def day_headers():
    return "".join(f'<th class="num">{d}</th>' for d in days)

def combined_rows():
    out = []
    for i, x in enumerate(combined):
        cls = "annule" if "Annul" in x["name"] else ""
        out.append(f'<tr class="{cls}"><td>{i+1}</td><td>{x["name"]}</td>'
                   f'<td class="num">{x["kr"]}</td><td class="num">{x["sm"]}</td>'
                   f'<td class="num"><strong>{x["total"]}</strong></td></tr>')
    return "\n".join(out)

pct_sm = round(total_sm/total_all*100) if total_all else 0
pct_kr = round(total_kr/total_all*100) if total_all else 0
today  = datetime.date.today().strftime("%d %B %Y")

js_days        = json.dumps(days)
js_top12_names = json.dumps([x["name"][:22] for x in top12])
js_top12_sm    = json.dumps([x["sm"]  for x in top12])
js_top12_kr    = json.dumps([x["kr"]  for x in top12])
js_wk1sm       = json.dumps(wk1_sm["grand"])
js_wk1kr       = json.dumps(wk1_kr["grand"])
js_wk2sm       = json.dumps(wk2_sm["grand"])
js_wk2kr       = json.dumps(wk2_kr["grand"])
js_wk1_daily   = json.dumps(wk1_daily)
js_wk2_daily   = json.dumps(wk2_daily)

CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',sans-serif;background:#0f1117;color:#e2e8f0;min-height:100vh;-webkit-text-size-adjust:100%}
.header{background:linear-gradient(135deg,#1a1f2e 0%,#16213e 50%,#0f3460 100%);padding:28px 16px;text-align:center;border-bottom:1px solid #2d3748}
.header h1{font-size:clamp(1.3rem,5vw,2rem);font-weight:700;color:#fff;letter-spacing:-0.5px}
.header p{color:#94a3b8;margin-top:8px;font-size:clamp(0.8rem,3vw,0.95rem)}
.badge{display:inline-block;background:rgba(99,102,241,0.2);border:1px solid rgba(99,102,241,0.4);color:#a5b4fc;padding:4px 14px;border-radius:20px;font-size:0.8rem;margin-top:12px}
.container{max-width:1400px;margin:0 auto;padding:20px 12px}
.kpi-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:24px}
.kpi{background:linear-gradient(135deg,#1e2535,#252d40);border:1px solid #2d3748;border-radius:14px;padding:18px 12px;text-align:center;transition:transform .2s}
.kpi:active{transform:scale(0.98)}
.kpi .val{font-size:clamp(1.6rem,6vw,2.8rem);font-weight:700;background:linear-gradient(135deg,#6366f1,#8b5cf6);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.kpi .lbl{color:#64748b;font-size:clamp(0.65rem,2vw,0.85rem);margin-top:6px;text-transform:uppercase;letter-spacing:.5px}
.kpi .sub{color:#94a3b8;font-size:clamp(0.75rem,2.5vw,0.9rem);margin-top:4px}
.section-title{font-size:clamp(0.95rem,3vw,1.1rem);font-weight:600;color:#c7d2fe;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #2d3748;display:flex;align-items:center;gap:8px}
.section-title::before{content:'';flex-shrink:0;width:4px;height:18px;background:linear-gradient(#6366f1,#8b5cf6);border-radius:2px}
.charts-2col{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px}
.chart-card{background:#1e2535;border:1px solid #2d3748;border-radius:14px;padding:18px 14px}
.chart-card h3{font-size:0.8rem;font-weight:600;color:#94a3b8;margin-bottom:14px;text-transform:uppercase;letter-spacing:.5px}
.chart-wrap{position:relative;height:240px}
.chart-wrap.tall{height:340px}
.chart-wrap.top12{height:300px}
table{width:100%;border-collapse:collapse;font-size:0.8rem}
th{background:#252d40;color:#94a3b8;padding:9px 8px;text-align:left;font-weight:500;text-transform:uppercase;font-size:0.68rem;letter-spacing:.4px;white-space:nowrap}
td{padding:7px 8px;border-bottom:1px solid #1a2030;color:#cbd5e1;white-space:nowrap}
tr:last-child td{border-bottom:none}
tr.total-row td{background:#252d40;color:#e2e8f0}
.num{text-align:right;font-variant-numeric:tabular-nums}
.annule{opacity:.5;font-style:italic}
.tbl-scroll{overflow-x:auto;-webkit-overflow-scrolling:touch;border-radius:8px}
.footer{text-align:center;padding:28px 16px;color:#475569;font-size:0.78rem;border-top:1px solid #1e2535}
/* Accordion */
.accordion{border:1px solid #2d3748;border-radius:14px;overflow:hidden;margin-bottom:16px}
.accordion-header{display:flex;align-items:center;justify-content:space-between;padding:16px 18px;background:#1e2535;cursor:pointer;user-select:none;-webkit-tap-highlight-color:transparent}
.accordion-header h3{font-size:0.85rem;font-weight:600;color:#c7d2fe;text-transform:uppercase;letter-spacing:.5px;margin:0}
.accordion-header .chevron{color:#6366f1;font-size:1.1rem;transition:transform .3s;flex-shrink:0}
.accordion-header.open .chevron{transform:rotate(180deg)}
.accordion-body{display:none;background:#1e2535;border-top:1px solid #2d3748}
.accordion-body.open{display:block}
.accordion-body .inner{padding:16px 14px}
/* Responsive breakpoints */
@media(max-width:700px){
  .kpi-grid{grid-template-columns:repeat(3,1fr);gap:8px}
  .kpi{padding:14px 8px}
  .charts-2col{grid-template-columns:1fr}
  .chart-wrap{height:160px}
  .chart-wrap.tall{height:200px}
  .chart-wrap.top12{height:260px}
  .container{padding:16px 10px}
  .chart-card{padding:14px 12px}
}
@media(max-width:400px){
  .kpi-grid{grid-template-columns:1fr 1fr;gap:8px}
  .kpi:last-child{grid-column:1/-1}
  .chart-wrap{height:140px}
  .chart-wrap.tall{height:180px}
  .chart-wrap.top12{height:220px}
}
"""

JS = f"""
const P='#6366f1', G='#10b981';
const grid='rgba(255,255,255,0.06)', tick='#64748b';
function baseOpts(horiz){{
  return {{
    responsive:true, maintainAspectRatio:false,
    indexAxis: horiz ? 'y' : 'x',
    plugins:{{
      legend:{{labels:{{color:'#94a3b8',font:{{size:11}}}}}},
      tooltip:{{backgroundColor:'#1e2535',borderColor:'#374151',borderWidth:1,titleColor:'#e2e8f0',bodyColor:'#94a3b8'}}
    }},
    scales:{{
      x:{{ticks:{{color:tick,font:{{size:10}}}},grid:{{color:grid}}}},
      y:{{ticks:{{color:tick,font:{{size:10}}}},grid:{{color:grid}},beginAtZero:true}}
    }}
  }};
}}

new Chart('cSplit',{{type:'doughnut',
  data:{{labels:['St Martin ({total_sm})','Kings Road ({total_kr})'],
    datasets:[{{data:[{total_sm},{total_kr}],backgroundColor:[G,P],borderWidth:0,hoverOffset:8}}]}},
  options:{{responsive:true,maintainAspectRatio:false,cutout:'65%',
    plugins:{{legend:{{position:'bottom',labels:{{color:'#94a3b8',padding:16}}}},
    tooltip:{{backgroundColor:'#1e2535',borderColor:'#374151',borderWidth:1,titleColor:'#e2e8f0',bodyColor:'#94a3b8'}}}}
  }}
}});

new Chart('cTop12',{{type:'bar',
  data:{{labels:{js_top12_names},
    datasets:[
      {{label:'Kings Road',data:{js_top12_kr},backgroundColor:P,borderRadius:4}},
      {{label:'St Martin', data:{js_top12_sm},backgroundColor:G,borderRadius:4}}
    ]}},
  options:baseOpts(true)
}});

new Chart('cW1SM',{{type:'bar',data:{{labels:{js_days},datasets:[{{label:'Items',data:{js_wk1sm},backgroundColor:G,borderRadius:6}}]}},options:baseOpts()}});
new Chart('cW1KR',{{type:'bar',data:{{labels:{js_days},datasets:[{{label:'Items',data:{js_wk1kr},backgroundColor:P,borderRadius:6}}]}},options:baseOpts()}});
new Chart('cW2SM',{{type:'bar',data:{{labels:{js_days},datasets:[{{label:'Items',data:{js_wk2sm},backgroundColor:G,borderRadius:6}}]}},options:baseOpts()}});
new Chart('cW2KR',{{type:'bar',data:{{labels:{js_days},datasets:[{{label:'Items',data:{js_wk2kr},backgroundColor:P,borderRadius:6}}]}},options:baseOpts()}});

new Chart('cTrend',{{type:'line',
  data:{{labels:{js_days},datasets:[
    {{label:'Week 1',data:{js_wk1_daily},borderColor:P,backgroundColor:'rgba(99,102,241,0.12)',fill:true,tension:0.4,pointRadius:5,pointBackgroundColor:P}},
    {{label:'Week 2',data:{js_wk2_daily},borderColor:G,backgroundColor:'rgba(16,185,129,0.12)',fill:true,tension:0.4,pointRadius:5,pointBackgroundColor:G}}
  ]}},
  options:baseOpts()
}});
"""

HTML = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>St George – Sales Report 04/05/26–17/05/26</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet"/>
<style>{CSS}</style>
</head>
<body>
<div class="header">
  <h1>&#x1F4CA; Weekly Sales Report</h1>
  <p>St George Restaurant Group &middot; Square POS Data</p>
  <span class="badge">04 May 2026 &ndash; 17 May 2026</span>
</div>
<div class="container">

  <div class="kpi-grid">
    <div class="kpi"><div class="val">{total_all}</div><div class="lbl">Total Items Sold</div><div class="sub">Both locations &middot; 2 weeks</div></div>
    <div class="kpi"><div class="val">{total_sm}</div><div class="lbl">St Martin</div><div class="sub">{pct_sm}% of total</div></div>
    <div class="kpi"><div class="val">{total_kr}</div><div class="lbl">Kings Road</div><div class="sub">{pct_kr}% of total</div></div>
  </div>

  <div class="charts-2col">
    <div class="chart-card"><h3>Location Split</h3><div class="chart-wrap"><canvas id="cSplit"></canvas></div></div>
    <div class="chart-card"><h3>Top 12 Items &ndash; Combined (Both Locations)</h3><div class="chart-wrap top12"><canvas id="cTop12"></canvas></div></div>
  </div>

  <div class="section-title">Combined Daily Trend &ndash; Both Locations</div>
  <div class="chart-card" style="margin-bottom:24px">
    <h3>Total Items Sold per Day</h3>
    <div class="chart-wrap"><canvas id="cTrend"></canvas></div>
  </div>

  <div class="section-title">Week 1 &middot; 04 May &ndash; 10 May 2026</div>
  <div class="charts-2col">
    <div class="chart-card"><h3>St Martin &ndash; Daily Sales</h3><div class="chart-wrap"><canvas id="cW1SM"></canvas></div></div>
    <div class="chart-card"><h3>Kings Road &ndash; Daily Sales</h3><div class="chart-wrap"><canvas id="cW1KR"></canvas></div></div>
  </div>

  <div class="accordion">
    <div class="accordion-header" onclick="toggleAcc(this)">
      <h3>St Martin &ndash; Week 1 Detail</h3><span class="chevron">&#x25BC;</span>
    </div>
    <div class="accordion-body">
      <div class="inner"><div class="tbl-scroll"><table>
        <thead><tr><th>Item</th>{day_headers()}<th class="num">Total</th></tr></thead>
        <tbody>{rows_html(wk1_sm["items"])}{total_row_html(wk1_sm["grand"],wk1_sm["grand_total"])}</tbody>
      </table></div></div>
    </div>
  </div>

  <div class="accordion">
    <div class="accordion-header" onclick="toggleAcc(this)">
      <h3>Kings Road &ndash; Week 1 Detail</h3><span class="chevron">&#x25BC;</span>
    </div>
    <div class="accordion-body">
      <div class="inner"><div class="tbl-scroll"><table>
        <thead><tr><th>Item</th>{day_headers()}<th class="num">Total</th></tr></thead>
        <tbody>{rows_html(wk1_kr["items"])}{total_row_html(wk1_kr["grand"],wk1_kr["grand_total"])}</tbody>
      </table></div></div>
    </div>
  </div>

  <div class="section-title" style="margin-top:24px">Week 2 &middot; 11 May &ndash; 17 May 2026</div>
  <div class="charts-2col">
    <div class="chart-card"><h3>St Martin &ndash; Daily Sales</h3><div class="chart-wrap"><canvas id="cW2SM"></canvas></div></div>
    <div class="chart-card"><h3>Kings Road &ndash; Daily Sales</h3><div class="chart-wrap"><canvas id="cW2KR"></canvas></div></div>
  </div>

  <div class="accordion">
    <div class="accordion-header" onclick="toggleAcc(this)">
      <h3>St Martin &ndash; Week 2 Detail</h3><span class="chevron">&#x25BC;</span>
    </div>
    <div class="accordion-body">
      <div class="inner"><div class="tbl-scroll"><table>
        <thead><tr><th>Item</th>{day_headers()}<th class="num">Total</th></tr></thead>
        <tbody>{rows_html(wk2_sm["items"])}{total_row_html(wk2_sm["grand"],wk2_sm["grand_total"])}</tbody>
      </table></div></div>
    </div>
  </div>

  <div class="accordion">
    <div class="accordion-header" onclick="toggleAcc(this)">
      <h3>Kings Road &ndash; Week 2 Detail</h3><span class="chevron">&#x25BC;</span>
    </div>
    <div class="accordion-body">
      <div class="inner"><div class="tbl-scroll"><table>
        <thead><tr><th>Item</th>{day_headers()}<th class="num">Total</th></tr></thead>
        <tbody>{rows_html(wk2_kr["items"])}{total_row_html(wk2_kr["grand"],wk2_kr["grand_total"])}</tbody>
      </table></div></div>
    </div>
  </div>

  <div class="section-title" style="margin-top:24px">All Items &ndash; Combined (04/05/26&ndash;17/05/26)</div>
  <div class="chart-card" style="margin-bottom:24px">
    <div class="tbl-scroll"><table>
      <thead><tr><th>#</th><th>Item</th><th class="num">Kings Road</th><th class="num">St Martin</th><th class="num">Total</th></tr></thead>
      <tbody>{combined_rows()}</tbody>
    </table></div>
  </div>

</div>
<div class="footer">Generated {today} &middot; St George Restaurant Group</div>
<script>
function toggleAcc(header){{
  header.classList.toggle('open');
  var body = header.nextElementSibling;
  body.classList.toggle('open');
}}
{JS}
</script>
</body>
</html>"""

out = os.path.join(BASE, "Sales_Report_04-17_May_2026.html")
with open(out, "w", encoding="utf-8") as f:
    f.write(HTML)
print("Done:", out)

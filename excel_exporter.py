"""
excel_exporter.py
=================
Writes all three reports into a professionally formatted Excel workbook.
Returns the workbook as an in-memory bytes buffer for Streamlit download.

Design
------
- Three separate sheets: 'Combined', 'Kings Road', 'St Martin'.
- Each sheet has a merged title row, bold + coloured header row, bordered data
  cells, auto-fitted column widths, and a bold total row.
- All styling is driven by constants defined at the top — change colours etc.
  here without touching logic.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants — edit these to change the look without touching logic
# ---------------------------------------------------------------------------

# Sheet tab colours (RGB hex without #)
SHEET_TAB_COLOURS = {
    "Combined":   "1F4E79",   # dark blue
    "Kings Road": "833C00",   # dark orange/brown
    "St Martin":  "375623",   # dark green
}

# Header row fill colours per sheet
HEADER_FILL_COLOURS = {
    "Combined":   "2E75B6",   # blue
    "Kings Road": "C55A11",   # orange
    "St Martin":  "538135",   # green
}

TITLE_FONT_SIZE = 13
HEADER_FONT_SIZE = 11
DATA_FONT_SIZE = 11
FONT_NAME = "Calibri"

THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)

TOTAL_ROW_FILL = "D9D9D9"   # light grey for the summary/total row


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _make_header_fill(colour_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour_hex)


def _write_report_to_sheet(
    ws,
    df: pd.DataFrame,
    title: str,
    sheet_key: str,
    start_row: int = 1,
    start_col: int = 1,
) -> None:
    """
    Write a single report DataFrame to *ws* starting at (*start_row*, *start_col*).

    Layout (rows relative to start_row)
    ------------------------------------
    start_row     : merged title
    start_row + 1 : column headers
    start_row + 2 onwards : data + total row
    """
    n_cols = len(df.columns)

    header_fill = _make_header_fill(HEADER_FILL_COLOURS[sheet_key])

    # ── Title row ─────────────────────────────────────────────────────────
    title_row = start_row
    ws.merge_cells(
        start_row=title_row, start_column=start_col,
        end_row=title_row,   end_column=start_col + n_cols - 1,
    )
    title_cell = ws.cell(row=title_row, column=start_col, value=title)
    title_cell.font = Font(
        name=FONT_NAME,
        bold=True,
        size=TITLE_FONT_SIZE,
        color="FFFFFF",
    )
    title_cell.fill = _make_header_fill(SHEET_TAB_COLOURS[sheet_key])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = THIN_BORDER
    ws.row_dimensions[title_row].height = 22

    # ── Header row ────────────────────────────────────────────────────────
    header_row = start_row + 1
    for col_offset, col_name in enumerate(df.columns):
        col_idx = start_col + col_offset
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(name=FONT_NAME, bold=True, size=HEADER_FONT_SIZE, color="FFFFFF")
        cell.fill = header_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center" if col_name in ("#SL",) else "left",
            vertical="center",
        )
    ws.row_dimensions[header_row].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────
    total_row_index = len(df) - 1   # last row in df is the total row (0-based)

    for df_row_idx, (_, row) in enumerate(df.iterrows()):
        ws_row = start_row + 2 + df_row_idx
        is_total = df_row_idx == total_row_index

        for col_offset, col_name in enumerate(df.columns):
            col_idx = start_col + col_offset
            value = row[col_name]
            # Coerce numpy types to native Python for openpyxl
            if hasattr(value, "item"):
                value = value.item()

            cell = ws.cell(row=ws_row, column=col_idx, value=value if value != "" else None)
            cell.border = THIN_BORDER
            cell.font = Font(name=FONT_NAME, size=DATA_FONT_SIZE, bold=is_total)

            if is_total:
                cell.fill = _make_header_fill(TOTAL_ROW_FILL)
                cell.font = Font(name=FONT_NAME, size=DATA_FONT_SIZE, bold=True)

            # Alignment
            if col_name == "#SL":
                cell.alignment = Alignment(horizontal="center")
            elif col_name == "Item Name":
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")

    # ── Auto-fit column widths ────────────────────────────────────────────
    for col_offset, col_name in enumerate(df.columns):
        col_letter = get_column_letter(start_col + col_offset)
        max_width = max(
            len(str(col_name)),
            *(
                len(str(row[col_name])) if str(row[col_name]) != "" else 0
                for _, row in df.iterrows()
            ),
        )
        ws.column_dimensions[col_letter].width = min(max_width + 4, 45)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

GAP_COLS = 1   # blank columns between reports

def build_excel(
    report1: pd.DataFrame, title1: str,
    report2: pd.DataFrame, title2: str,
    report3: pd.DataFrame, title3: str,
) -> bytes:
    """
    Write all three reports side by side on a single worksheet.

    Layout
    ------
    [Combined] [gap] [Kings Road] [gap] [St Martin]
    All starting at row 1, each offset by its column width + GAP_COLS.

    Returns
    -------
    bytes : the .xlsx file as bytes (for st.download_button).
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="Sales Report")
    ws.sheet_properties.tabColor = SHEET_TAB_COLOURS["Combined"]

    reports = [
        (report1, title1, "Combined"),
        (report2, title2, "Kings Road"),
        (report3, title3, "St Martin"),
    ]

    current_col = 1
    for df, title, sheet_key in reports:
        _write_report_to_sheet(ws, df, title, sheet_key, start_row=1, start_col=current_col)
        # advance by this report’s column count + gap
        current_col += len(df.columns) + GAP_COLS

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

